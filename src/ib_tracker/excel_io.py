"""Excel tracker: workbook layout, styling, and read/write helpers.

The tracker workbook has three sheets, each with a title banner in row 1, a
status/summary banner in row 2 (a single writable cell anchoring a merged
range), a header row in row 3, and data starting at row 4:

  - "Master Job Tracker": one row per open application/opportunity.
  - "Bank Coverage Universe": one row per bank, updated on every run with
    when it was last checked and what was found.
  - "Search Log": one row per (bank, run) with what that check turned up.
"""

from __future__ import annotations

import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .banks import BankSource, ManualBank

NAVY = "1F3864"
LIGHT_GRN = "E2EFDA"
LIGHT_BLUE = "D6E4F7"
WHITE = "FFFFFF"

MASTER_COLS = [
    "#", "Bank Name", "Bank Category", "Division / Group", "Job Title", "Job Type",
    "Class Year", "Location", "Office / City", "Region", "Posting Status", "Priority",
    "Fit Score", "Application Link", "Careers Website", "Req ID", "Date First Found",
    "Date Last Checked", "Posting Date", "App Deadline", "Est. Deadline",
    "Days Since Posted", "Notes on Role", "Eligibility Notes",
    "Sponsorship / Work Auth", "Resume Version", "Cover Letter?", "Transcript?",
    "Networking Contact", "Contact Email / LinkedIn", "Last Touch", "Follow-Up?",
    "App Submitted Date", "Interview Status", "Next Step", "Final Outcome", "Notes",
]

# Bank Name must land at index 1 (column B) and Last Checked Date / Current
# Status at indices 9/10 (columns J/K) — update_coverage_universe() below
# addresses rows positionally to match that.
BANK_UNIV_COLS = [
    "#", "Bank Name", "Category", "ATS Platform", "Careers URL", "Priority",
    "Class Year Focus", "Manual-Check Note", "Notes", "Last Checked Date",
    "Current Status",
]

SEARCH_LOG_COLS = [
    "#", "Date Checked", "Time", "Bank", "Sample Link", "Search Terms",
    "Found?", "# Postings", "Job Titles", "Links", "Source", "Priority", "Notes",
]

HEADER_ROW = 3
FIRST_DATA_ROW = 4

MASTER_TABLE_NAME = "JobTracker"

# Sort order for the Master Job Tracker's default view — "job type" groups
# together, Investment Banking first since that's the primary target.
DIVISION_SORT_ORDER = {
    "Investment Banking": 0,
    "Markets / Sales & Trading": 1,
    "Research": 2,
    "FT Analyst Program (verify division)": 3,
}


def _division_sort_key(division: str) -> int:
    return DIVISION_SORT_ORDER.get(division, 99)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def get_column_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter."""
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def col_index(name: str, cols: list[str] = MASTER_COLS) -> int:
    return cols.index(name) + 1


def _write_header_row(ws: Worksheet, cols: list[str], row: int = HEADER_ROW) -> None:
    for ci, name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.border = _border()
        cell.alignment = _center()
    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def _apply_table(ws: Worksheet, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    """Turn a header+data range into a real Excel Table: gives it filter/sort
    dropdown arrows on the header row and banded rows, in Excel, Google
    Sheets, and LibreOffice alike."""
    if name in ws.tables:
        del ws.tables[name]
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(table)


def _set_autofilter(ws: Worksheet, ref: str) -> None:
    ws.auto_filter.ref = ref


def _write_title_and_banner(ws: Worksheet, title: str, banner: str, n_cols: int) -> None:
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = banner
    ws["A2"].font = Font(name="Arial", size=9, italic=True)
    ws["A2"].fill = _fill(LIGHT_BLUE)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16


def build_tracker_workbook(class_year: str, banks: list[BankSource], manual_banks: list[ManualBank]) -> Workbook:
    """Create a fresh tracker workbook: headers, styling, and the bank
    universe pre-populated from the registry. No job rows yet — those get
    added by the first `update` run."""
    wb = Workbook()

    ws_master = wb.active
    ws_master.title = "Master Job Tracker"
    _write_title_and_banner(
        ws_master,
        f"Master Job Tracker — Class of {class_year} Investment Banking Full-Time Analyst Recruiting",
        f"Tracker initialized: {datetime.date.today():%Y-%m-%d}",
        len(MASTER_COLS),
    )
    _write_header_row(ws_master, MASTER_COLS)
    for ci, name in enumerate(MASTER_COLS, 1):
        ws_master.column_dimensions[get_column_letter(ci)].width = 16 if name != "Notes" else 30
    _apply_table(ws_master, MASTER_TABLE_NAME, f"A{HEADER_ROW}:{get_column_letter(len(MASTER_COLS))}{HEADER_ROW}")

    ws_univ = wb.create_sheet("Bank Coverage Universe")
    _write_title_and_banner(ws_univ, "Bank Coverage Universe", "Refreshed on every `ib-tracker update` run.", len(BANK_UNIV_COLS))
    _write_header_row(ws_univ, BANK_UNIV_COLS)
    row = FIRST_DATA_ROW
    for seq, bank in enumerate(banks, 1):
        _write_coverage_row(ws_univ, row, seq, bank.name, bank.category, bank.platform,
                             bank.careers_url, "High", class_year, "", "", "", "Not yet checked")
        row += 1
    for seq, mb in enumerate(manual_banks, len(banks) + 1):
        _write_coverage_row(ws_univ, row, seq, mb.name, mb.category, "Manual",
                             mb.careers_url, "Medium", class_year, mb.note, "", "", "Check manually — no public API")
        row += 1
    for ci, _name in enumerate(BANK_UNIV_COLS, 1):
        ws_univ.column_dimensions[get_column_letter(ci)].width = 18
    _set_autofilter(ws_univ, f"A{HEADER_ROW}:{get_column_letter(len(BANK_UNIV_COLS))}{row - 1}")

    ws_log = wb.create_sheet("Search Log")
    _write_title_and_banner(ws_log, "Search Log", "One row per bank per `ib-tracker update` run.", len(SEARCH_LOG_COLS))
    _write_header_row(ws_log, SEARCH_LOG_COLS)
    for ci, _name in enumerate(SEARCH_LOG_COLS, 1):
        ws_log.column_dimensions[get_column_letter(ci)].width = 20
    _set_autofilter(ws_log, f"A{HEADER_ROW}:{get_column_letter(len(SEARCH_LOG_COLS))}{HEADER_ROW}")

    return wb


def _write_coverage_row(ws, row, seq, bank_name, category, platform, careers_url,
                         priority, class_year, manual_note, notes, last_checked, status) -> None:
    values = [seq, bank_name, category, platform, careers_url, priority, class_year,
              manual_note, notes, last_checked, status]
    alt = row % 2 == 0
    bg = "EBF3FF" if alt else WHITE
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill = _fill(bg)
        cell.border = _border()
        cell.font = Font(name="Arial", size=9)
        cell.alignment = _left()


def existing_links(ws: Worksheet) -> set[str]:
    """Return the set of application links already in the Master Job Tracker sheet."""
    links = set()
    link_col = col_index("Application Link")
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        v = row[link_col - 1]
        if v:
            links.add(str(v).strip())
    return links


def next_empty_row(ws: Worksheet) -> int:
    link_col = col_index("Application Link")
    for row_idx in range(FIRST_DATA_ROW, ws.max_row + 2):
        if ws.cell(row_idx, link_col).value in (None, ""):
            return row_idx
    return ws.max_row + 1


def write_job_row(ws: Worksheet, row: int, job: dict, seq: int) -> None:
    today = datetime.date.today().strftime("%m/%d/%Y")

    values = {
        "#": seq,
        "Bank Name": job["bank"],
        "Bank Category": job.get("category", ""),
        "Division / Group": job.get("division", ""),
        "Job Title": job["title"],
        "Job Type": "Full-Time Analyst",
        "Location": job.get("loc", ""),
        "Posting Status": "Open",
        "Priority": "High",
        "Application Link": job["link"],
        "Date First Found": today,
        "Date Last Checked": today,
    }

    for ci in range(1, len(MASTER_COLS) + 1):
        col_name = MASTER_COLS[ci - 1]
        val = values.get(col_name, "")
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill = _fill(LIGHT_GRN)
        cell.border = _border()
        cell.alignment = _center() if col_name in (
            "#", "Posting Status", "Priority", "Job Type", "Date First Found", "Date Last Checked"
        ) else _left()
        cell.font = Font(name="Arial", size=9, bold=True)

    dsp_col = col_index("Days Since Posted")
    posting_col = get_column_letter(col_index("Posting Date"))
    dsp_cell = ws.cell(row=row, column=dsp_col)
    dsp_cell.value = f'=IF({posting_col}{row}="","",TODAY()-{posting_col}{row})'
    dsp_cell.number_format = "0"

    ws.row_dimensions[row].height = 18


def finalize_master_sheet(ws: Worksheet, new_links: set[str] | None = None) -> None:
    """Re-sort every job row by division ("job type" — Investment Banking,
    then Markets, Research, unclear), renumber them, restyle them (only
    rows newly found in this run get the highlight), and (re)apply the
    Excel Table over the full range so the header row's filter/sort
    dropdowns cover every row. Call this once per `update` run, after all
    per-bank rows have been appended."""
    new_links = new_links or set()
    n_cols = len(MASTER_COLS)
    link_idx = col_index("Application Link") - 1
    division_idx = col_index("Division / Group") - 1
    bank_idx = col_index("Bank Name") - 1
    title_idx = col_index("Job Title") - 1

    rows = [
        list(row)
        for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=ws.max_row, max_col=n_cols, values_only=True)
        if row[link_idx]
    ]
    rows.sort(key=lambda r: (_division_sort_key(r[division_idx] or ""), r[bank_idx] or "", r[title_idx] or ""))

    last_row = max(ws.max_row, FIRST_DATA_ROW + len(rows) - 1) if rows else ws.max_row
    for r in range(FIRST_DATA_ROW, last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(r, c).value = None

    posting_col = get_column_letter(col_index("Posting Date"))
    dsp_col = col_index("Days Since Posted")
    seq_col = col_index("#")

    for i, values in enumerate(rows):
        row_num = FIRST_DATA_ROW + i
        is_new = values[link_idx] in new_links
        for ci, val in enumerate(values, 1):
            val = (i + 1) if ci == seq_col else val
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = _border()
            col_name = MASTER_COLS[ci - 1]
            cell.alignment = _center() if col_name in (
                "#", "Posting Status", "Priority", "Job Type", "Date First Found", "Date Last Checked"
            ) else _left()
            if is_new:
                cell.fill = _fill(LIGHT_GRN)
                cell.font = Font(name="Arial", size=9, bold=True)
            else:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(name="Arial", size=9)
        ws.cell(row_num, dsp_col).value = f'=IF({posting_col}{row_num}="","",TODAY()-{posting_col}{row_num})'
        ws.cell(row_num, dsp_col).number_format = "0"
        ws.row_dimensions[row_num].height = 18

    table_last_row = FIRST_DATA_ROW + len(rows) - 1 if rows else HEADER_ROW
    _apply_table(ws, MASTER_TABLE_NAME, f"A{HEADER_ROW}:{get_column_letter(n_cols)}{table_last_row}")


def refresh_search_log_filter(ws_log: Worksheet) -> None:
    """Search Log grows every run — widen its autofilter to cover all rows."""
    _set_autofilter(ws_log, f"A{HEADER_ROW}:{get_column_letter(len(SEARCH_LOG_COLS))}{ws_log.max_row}")


def update_search_log(ws_log: Worksheet, bank_name: str, jobs: list[dict], checked_at: str) -> None:
    """Append a row to the Search Log tab."""
    next_row = ws_log.max_row + 1
    for r in range(FIRST_DATA_ROW, ws_log.max_row + 2):
        if ws_log.cell(r, 2).value in (None, ""):
            next_row = r
            break

    bg = "EBF3FF" if next_row % 2 == 0 else WHITE
    found = "Yes" if jobs else "No"
    titles = "; ".join(j["title"] for j in jobs) if jobs else "—"
    links = "; ".join(j["link"] for j in jobs) if jobs else "—"

    row_vals = [
        next_row - (FIRST_DATA_ROW - 1),
        checked_at,
        datetime.datetime.now().strftime("%H:%M"),
        bank_name,
        jobs[0].get("link", "") if jobs else "—",
        "Investment Banking Analyst, Full-Time Analyst",
        found,
        len(jobs),
        titles,
        links,
        jobs[0].get("source", "Direct Website") if jobs else "Direct Website",
        "High",
        "",
    ]
    for ci, val in enumerate(row_vals, 1):
        cell = ws_log.cell(row=next_row, column=ci, value=val)
        cell.fill = _fill(bg)
        cell.font = Font(name="Arial", size=9)
        cell.border = _border()
        cell.alignment = _left()
    ws_log.row_dimensions[next_row].height = 16


def update_coverage_universe(ws_univ: Worksheet, bank_name: str, status: str, checked_at: str) -> None:
    """Update Last Checked Date and Current Status in the Bank Coverage Universe tab."""
    for row in ws_univ.iter_rows(min_row=FIRST_DATA_ROW):
        if str(row[1].value).strip().lower() == bank_name.strip().lower():
            row[9].value = checked_at
            row[10].value = status
            row[9].number_format = "MM/DD/YYYY"
            break
