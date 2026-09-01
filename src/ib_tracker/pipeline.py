"""High-level operations that tie fetchers, banks, and the Excel tracker
together: fetching every bank's postings, and the two things you do with the
results — a quick console report (`check`) or an Excel update (`update`)."""

from __future__ import annotations

import datetime
import json
import re
import sys
import time
from pathlib import Path
from typing import Callable

import requests
from openpyxl import load_workbook

from . import excel_io
from .banks import BANKS, NO_PUBLIC_API, BankSource
from .config import BANK_PAUSE_SECONDS, HEADERS, Config
from .fetchers import FetchContext
from .location import is_us_posting


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def fetch_all(
    cfg: Config,
    session: requests.Session | None = None,
    *,
    on_bank_done: Callable[[BankSource, list[dict]], None] | None = None,
) -> list[tuple[BankSource, list[dict]]]:
    """Query every bank in the registry. Returns [(bank, jobs), ...] in
    registry order. Each job dict is tagged with its bank's category, and
    non-US postings are filtered out (see location.is_us_posting)."""
    session = session or build_session()
    ctx = FetchContext(class_year=cfg.class_year, search_terms=cfg.search_terms, session=session)
    results = []
    for bank in BANKS:
        jobs = bank.fetch(ctx, bank.name, **bank.kwargs)
        jobs = [j for j in jobs if is_us_posting(j.get("loc", ""), j.get("title", ""))]
        for j in jobs:
            j["category"] = bank.category
        if on_bank_done:
            on_bank_done(bank, jobs)
        results.append((bank, jobs))
        time.sleep(BANK_PAUSE_SECONDS)
    return results


# ══════════════════════════════════════════════════════════════════════════
# `check` — console-only preview, no Excel file needed.
# ══════════════════════════════════════════════════════════════════════════

def run_check(cfg: Config, *, show_laterals: bool = False, stream=sys.stdout) -> int:
    """Print which class-year full-time programs are live, and what's new
    since the last check (tracked in a small seen-links JSON file). Returns
    the number of class-year drops found."""
    print(f"\n{'=' * 72}", file=stream)
    print(f"  IB FT Application Checker — Class of {cfg.class_year}", file=stream)
    print(f"  {datetime.datetime.now():%Y-%m-%d %H:%M}", file=stream)
    print(f"{'=' * 72}\n", file=stream)

    seen: set[str] = set()
    if cfg.seen_path.exists():
        try:
            seen = set(json.loads(cfg.seen_path.read_text()))
        except Exception:
            pass
    first_run = not seen

    def on_bank_done(bank: BankSource, jobs: list[dict]) -> None:
        print(f"  checking {bank.name:<24}{len(jobs)} match(es)", file=stream)

    all_jobs = [j for _bank, jobs in fetch_all(cfg, on_bank_done=on_bank_done) for j in jobs]

    def is_class_year_ft(j: dict) -> bool:
        return cfg.class_year in j["title"]

    drops = [j for j in all_jobs if is_class_year_ft(j)]
    laterals = [j for j in all_jobs if not is_class_year_ft(j)]
    new_links = {j["link"] for j in all_jobs} - seen

    print(f"\n{'─' * 72}", file=stream)
    print(f"  CLASS OF {cfg.class_year} FULL-TIME PROGRAMS LIVE RIGHT NOW: {len(drops)}", file=stream)
    print(f"{'─' * 72}", file=stream)
    for j in sorted(drops, key=lambda x: (x["link"] not in new_links, x["bank"])):
        flag = "NEW" if (j["link"] in new_links and not first_run) else "   "
        print(f"\n  {flag} {j['bank']} — {j['division']}", file=stream)
        print(f"        {j['title']}", file=stream)
        if j["loc"]:
            print(f"        {j['loc']}", file=stream)
        print(f"        {j['link']}", file=stream)

    if not drops:
        print(f"\n  none found yet — the {cfg.class_year} cycle may not have opened. Keep checking!", file=stream)

    if show_laterals and laterals:
        print(f"\n{'─' * 72}", file=stream)
        print(f"  other analyst matches (mostly experienced/lateral — verify before applying): {len(laterals)}", file=stream)
        for j in laterals:
            print(f"    {j['bank']:<22} | {j['division']:<28} | {j['title'][:70]}", file=stream)

    print(f"\n{'─' * 72}", file=stream)
    print(f"  NO PUBLIC FEED — CHECK THESE BY HAND:", file=stream)
    for mb in NO_PUBLIC_API:
        print(f"    {mb.name:<22} {mb.careers_url}", file=stream)
    print(file=stream)

    cfg.seen_path.write_text(json.dumps(sorted({j["link"] for j in all_jobs} | seen)))
    if first_run:
        print("  (first run — everything is baseline; next run flags NEW drops)\n", file=stream)

    return len(drops)


# ══════════════════════════════════════════════════════════════════════════
# `update` — writes new openings into the Excel tracker workbook.
# ══════════════════════════════════════════════════════════════════════════

def run_update(cfg: Config, *, dry_run: bool = False, stream=sys.stdout) -> list[dict]:
    """Fetch every bank, diff against what's already in the tracker, and
    (unless dry_run) write new rows into the Master Job Tracker sheet plus
    a status row for every bank in the Bank Coverage Universe and Search Log
    sheets. Returns the list of newly found job dicts."""
    if not cfg.tracker_path.exists():
        print(f"[ERROR] Tracker not found at {cfg.tracker_path} — run `ib-tracker init` first.", file=stream)
        sys.exit(1)

    print(f"\n{'=' * 65}", file=stream)
    print(f"  IB Recruiting Tracker Auto-Updater", file=stream)
    print(f"  {datetime.datetime.now():%Y-%m-%d %H:%M:%S}", file=stream)
    print(f"  Mode: {'DRY RUN (no writes)' if dry_run else 'LIVE UPDATE'}", file=stream)
    print(f"{'=' * 65}\n", file=stream)

    wb = load_workbook(cfg.tracker_path)
    ws_master = wb["Master Job Tracker"]
    ws_univ = wb["Bank Coverage Universe"]
    ws_log = wb["Search Log"]

    known_links = excel_io.existing_links(ws_master)
    today_str = datetime.date.today().strftime("%m/%d/%Y")

    seq_candidates = [
        ws_master.cell(r, 1).value
        for r in range(excel_io.FIRST_DATA_ROW, ws_master.max_row + 1)
    ]
    seq = max((v for v in seq_candidates if isinstance(v, int)), default=0) + 1

    all_new: list[dict] = []

    def on_bank_done(bank: BankSource, jobs: list[dict]) -> None:
        nonlocal seq
        new_jobs = [j for j in jobs if j.get("link", "") not in known_links]
        status = "Checked - Opening Found" if new_jobs else "Checked - No Opening"
        print(f"  Checking {bank.name:<35}", end="", file=stream)
        print(f"→ {len(new_jobs)} new opening(s)" if new_jobs else "→ No new openings", file=stream)

        if not dry_run:
            excel_io.update_coverage_universe(ws_univ, bank.name, status, today_str)
            excel_io.update_search_log(ws_log, bank.name, new_jobs, today_str)

        for job in new_jobs:
            all_new.append(job)
            if not dry_run:
                nr = excel_io.next_empty_row(ws_master)
                excel_io.write_job_row(ws_master, nr, job, seq)
                known_links.add(job.get("link", ""))
                seq += 1

    fetch_all(cfg, on_bank_done=on_bank_done)

    for mb in NO_PUBLIC_API:
        print(f"  Checking {mb.name:<35}→ Skipped ({mb.note})", file=stream)
        if not dry_run:
            excel_io.update_coverage_universe(ws_univ, mb.name, mb.note, today_str)

    if not dry_run:
        excel_io.finalize_master_sheet(ws_master, new_links={j["link"] for j in all_new})
        excel_io.refresh_search_log_filter(ws_log)

    if not dry_run and all_new:
        base_text = re.sub(r"\s*\|\s*Last auto-updated:.*$", "", ws_master["A2"].value or "")
        ws_master["A2"] = (
            f"{base_text}  |  Last auto-updated: {datetime.datetime.now():%Y-%m-%d %H:%M}"
            f"  |  {len(all_new)} new role(s) added"
        )

    if not dry_run:
        wb.save(cfg.tracker_path)
        print(f"\nTracker saved → {cfg.tracker_path}", file=stream)
    else:
        print(f"\n[DRY RUN] Would have added {len(all_new)} role(s).", file=stream)

    _append_log(cfg.log_path, dry_run, all_new)
    _print_update_summary(all_new, len(BANKS), len(NO_PUBLIC_API), stream)
    return all_new


def _append_log(log_path: Path, dry_run: bool, all_new: list[dict], *, keep_last: int = 90) -> None:
    log_entry = {
        "run_at": datetime.datetime.now().isoformat(),
        "dry_run": dry_run,
        "new_roles": len(all_new),
        "roles": all_new,
    }
    existing_log = []
    if log_path.exists():
        try:
            existing_log = json.loads(log_path.read_text())
        except Exception:
            pass
    existing_log.append(log_entry)
    log_path.write_text(json.dumps(existing_log[-keep_last:], indent=2))


def _print_update_summary(all_new: list[dict], n_banks: int, n_manual: int, stream) -> None:
    print(f"\n{'─' * 65}", file=stream)
    print(
        f"  SUMMARY: {len(all_new)} new opening(s) found across {n_banks} banks "
        f"({n_manual} more skipped — no public API, check manually)",
        file=stream,
    )
    print(f"{'─' * 65}", file=stream)
    if all_new:
        for j in all_new:
            print(f"\n  NEW OPENING", file=stream)
            print(f"     Bank    : {j['bank']} ({j.get('category', '')})", file=stream)
            print(f"     Division: {j.get('division', '')}", file=stream)
            print(f"     Title   : {j['title']}", file=stream)
            print(f"     Link    : {j['link']}", file=stream)
            print(f"     Source  : {j.get('source', '')}", file=stream)
    else:
        print("  No new openings found this run. Keep monitoring — roles can open any day.", file=stream)
    print(file=stream)


# ══════════════════════════════════════════════════════════════════════════
# `init` — create a fresh tracker workbook.
# ══════════════════════════════════════════════════════════════════════════

def run_init(cfg: Config, *, force: bool = False, stream=sys.stdout) -> None:
    if cfg.tracker_path.exists() and not force:
        print(f"[ERROR] {cfg.tracker_path} already exists. Pass --force to overwrite.", file=stream)
        sys.exit(1)
    wb = excel_io.build_tracker_workbook(cfg.class_year, BANKS, NO_PUBLIC_API)
    wb.save(cfg.tracker_path)
    print(f"Created tracker → {cfg.tracker_path}", file=stream)
