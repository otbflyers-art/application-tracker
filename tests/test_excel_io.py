from openpyxl import load_workbook

from ib_tracker import excel_io
from ib_tracker.banks import BankSource, ManualBank
from ib_tracker.fetchers import fetch_workday


def _sample_banks():
    banks = [
        BankSource("Test Bank", "Bulge Bracket", "Workday", fetch_workday,
                   {"tenant": "tb", "wd_num": 1, "site": "External"}, "https://example.com/careers"),
    ]
    manual = [ManualBank("No API Bank", "Elite Boutique", "No public API", "https://example.com/no-api")]
    return banks, manual


def test_build_tracker_workbook_has_expected_sheets_and_headers():
    banks, manual = _sample_banks()
    wb = excel_io.build_tracker_workbook("2027", banks, manual)

    assert wb.sheetnames == ["Master Job Tracker", "Bank Coverage Universe", "Search Log"]

    ws_master = wb["Master Job Tracker"]
    header = [c.value for c in ws_master[excel_io.HEADER_ROW]]
    assert header == excel_io.MASTER_COLS
    # A real Excel Table gives the header row filter/sort dropdown arrows.
    assert excel_io.MASTER_TABLE_NAME in ws_master.tables

    ws_univ = wb["Bank Coverage Universe"]
    header = [c.value for c in ws_univ[excel_io.HEADER_ROW]]
    assert header == excel_io.BANK_UNIV_COLS
    # Bank rows are pre-populated: one automated bank, one manual-check bank.
    bank_names = [row[1].value for row in ws_univ.iter_rows(min_row=excel_io.FIRST_DATA_ROW) if row[1].value]
    assert bank_names == ["Test Bank", "No API Bank"]
    assert ws_univ.auto_filter.ref is not None

    ws_log = wb["Search Log"]
    header = [c.value for c in ws_log[excel_io.HEADER_ROW]]
    assert header == excel_io.SEARCH_LOG_COLS
    assert ws_log.auto_filter.ref is not None


def test_write_job_row_and_existing_links_roundtrip(tmp_path):
    banks, manual = _sample_banks()
    wb = excel_io.build_tracker_workbook("2027", banks, manual)
    ws_master = wb["Master Job Tracker"]

    job = {
        "bank": "Test Bank",
        "title": "Investment Banking Analyst",
        "division": "Investment Banking",
        "link": "https://example.com/jobs/1",
        "category": "Bulge Bracket",
    }
    row = excel_io.next_empty_row(ws_master)
    assert row == excel_io.FIRST_DATA_ROW
    excel_io.write_job_row(ws_master, row, job, seq=1)

    assert excel_io.existing_links(ws_master) == {"https://example.com/jobs/1"}
    assert ws_master.cell(row, excel_io.col_index("Job Title")).value == "Investment Banking Analyst"
    assert ws_master.cell(row, excel_io.col_index("Posting Status")).value == "Open"

    next_row = excel_io.next_empty_row(ws_master)
    assert next_row == row + 1

    path = tmp_path / "tracker.xlsx"
    wb.save(path)
    reloaded = load_workbook(path)
    assert excel_io.existing_links(reloaded["Master Job Tracker"]) == {"https://example.com/jobs/1"}


def test_update_coverage_universe_and_search_log(tmp_path):
    banks, manual = _sample_banks()
    wb = excel_io.build_tracker_workbook("2027", banks, manual)
    ws_univ = wb["Bank Coverage Universe"]
    ws_log = wb["Search Log"]

    jobs = [{
        "bank": "Test Bank",
        "title": "Investment Banking Analyst",
        "link": "https://example.com/jobs/1",
        "source": "Workday",
    }]
    excel_io.update_coverage_universe(ws_univ, "Test Bank", "Checked - Opening Found", "01/01/2027")
    excel_io.update_search_log(ws_log, "Test Bank", jobs, "01/01/2027")

    for row in ws_univ.iter_rows(min_row=excel_io.FIRST_DATA_ROW):
        if row[1].value == "Test Bank":
            assert row[9].value == "01/01/2027"
            assert row[10].value == "Checked - Opening Found"
            break
    else:
        raise AssertionError("Test Bank row not found")

    log_row = ws_log[excel_io.FIRST_DATA_ROW]
    values = [c.value for c in log_row]
    assert values[3] == "Test Bank"
    assert values[6] == "Yes"
    assert values[7] == 1
    assert "Investment Banking Analyst" in values[8]


def test_finalize_master_sheet_sorts_by_division_and_highlights_new_rows():
    banks, manual = _sample_banks()
    wb = excel_io.build_tracker_workbook("2027", banks, manual)
    ws_master = wb["Master Job Tracker"]

    # Written out of order and out of sequence — finalize should fix both.
    jobs = [
        {"bank": "Test Bank", "title": "Research Analyst", "division": "Research",
         "link": "https://example.com/jobs/research", "category": "Bulge Bracket"},
        {"bank": "Test Bank", "title": "Investment Banking Analyst", "division": "Investment Banking",
         "link": "https://example.com/jobs/ib", "category": "Bulge Bracket"},
        {"bank": "Test Bank", "title": "Markets Analyst", "division": "Markets / Sales & Trading",
         "link": "https://example.com/jobs/markets", "category": "Bulge Bracket"},
    ]
    for i, job in enumerate(jobs, 1):
        row = excel_io.next_empty_row(ws_master)
        excel_io.write_job_row(ws_master, row, job, seq=i)

    # Only the IB and Markets postings are "new" in this run.
    excel_io.finalize_master_sheet(ws_master, new_links={
        "https://example.com/jobs/ib", "https://example.com/jobs/markets",
    })

    divisions = [
        row[excel_io.col_index("Division / Group") - 1]
        for row in ws_master.iter_rows(min_row=excel_io.FIRST_DATA_ROW, values_only=True)
        if row[excel_io.col_index("Application Link") - 1]
    ]
    assert divisions == ["Investment Banking", "Markets / Sales & Trading", "Research"]

    seqs = [
        row[excel_io.col_index("#") - 1]
        for row in ws_master.iter_rows(min_row=excel_io.FIRST_DATA_ROW, values_only=True)
        if row[excel_io.col_index("Application Link") - 1]
    ]
    assert seqs == [1, 2, 3]

    def _fill_color(row_offset, col_name):
        cell = ws_master.cell(excel_io.FIRST_DATA_ROW + row_offset, excel_io.col_index(col_name))
        return cell.fill.fgColor.rgb

    assert _fill_color(0, "Job Title") == "00" + excel_io.LIGHT_GRN  # Investment Banking row is new
    assert _fill_color(1, "Job Title") == "00" + excel_io.LIGHT_GRN  # Markets row is new
    assert _fill_color(2, "Job Title") != "00" + excel_io.LIGHT_GRN  # Research row is not new this run

    # The Table's range grows to cover all three rows.
    table = ws_master.tables[excel_io.MASTER_TABLE_NAME]
    assert table.ref == f"A{excel_io.HEADER_ROW}:{excel_io.get_column_letter(len(excel_io.MASTER_COLS))}{excel_io.FIRST_DATA_ROW + 2}"
